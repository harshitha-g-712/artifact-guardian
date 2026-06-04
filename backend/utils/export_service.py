"""
export_service.py  —  PDF (reportlab) and Excel (openpyxl) export + CSV import
"""
import io
import csv
from datetime import datetime


# ── Excel Export ──────────────────────────────────────────────────────────────

def export_inspections_excel(inspections: list) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise RuntimeError("openpyxl not installed — run: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inspections"

    # Header row
    headers = ["ID", "Artifact", "Date", "Type", "Inspector",
               "Crack Detected", "Fading Level", "Severity Index", "Risk", "Notes"]
    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    hdr_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    thin = Side(style="thin", color="C7D2FE")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 20

    col_widths = [6, 28, 12, 14, 18, 14, 12, 14, 10, 40]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    fill_even = PatternFill("solid", fgColor="F0F4FF")
    fill_odd  = PatternFill("solid", fgColor="FFFFFF")

    for ri, insp in enumerate(inspections, 2):
        sev = float(insp.get("severity_index", 0))
        risk = ("CRITICAL" if sev >= 8 else "HIGH" if sev >= 6 else "MEDIUM" if sev >= 3.5 else "LOW")
        row_data = [
            insp.get("inspection_id"),
            insp.get("artifact_name", ""),
            str(insp.get("inspection_date", ""))[:10],
            insp.get("inspection_type", "Routine"),
            insp.get("inspector_name", ""),
            "Yes" if insp.get("crack_detected") else "No",
            f"{float(insp.get('fading_level', 0)) * 100:.1f}%",
            f"{sev:.1f}",
            risk,
            insp.get("damage_notes", ""),
        ]
        fill = fill_even if ri % 2 == 0 else fill_odd
        for col, v in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=col, value=v)
            cell.fill = fill
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 10))

    # Freeze header
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── PDF Export ────────────────────────────────────────────────────────────────

def export_report_pdf(artifact: dict, inspections: list) -> bytes:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                        Table, TableStyle, HRFlowable)
    except ImportError:
        raise RuntimeError("reportlab not installed — run: pip install reportlab")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    navy = colors.HexColor("#1E3A5F")
    gold = colors.HexColor("#D4A017")
    story = []

    # ── Title ─────────────────────────────────────────
    title_style = ParagraphStyle("AG_Title", fontSize=24, textColor=navy,
                                 fontName="Helvetica-Bold", spaceAfter=4, alignment=TA_CENTER)
    sub_style = ParagraphStyle("AG_Sub", fontSize=12, textColor=colors.grey,
                               spaceAfter=4, alignment=TA_CENTER)
    date_style = ParagraphStyle("AG_Date", fontSize=10, textColor=colors.grey,
                                spaceAfter=20, alignment=TA_CENTER)

    story.append(Paragraph("ARTIFACT GUARDIAN", title_style))
    story.append(Paragraph("AI Heritage Preservation System", sub_style))
    story.append(Paragraph(f"Report generated: {datetime.now().strftime('%d %B %Y, %H:%M')}", date_style))
    story.append(HRFlowable(width="100%", thickness=2, color=navy, spaceAfter=16))

    # ── Artifact Details ──────────────────────────────
    h2 = ParagraphStyle("AG_H2", fontSize=14, textColor=navy,
                        fontName="Helvetica-Bold", spaceAfter=8, spaceBefore=12)
    story.append(Paragraph("Artifact Details", h2))

    sev_label = "CRITICAL" if (artifact.get("max_severity") or 0) >= 8 else \
                "HIGH" if (artifact.get("max_severity") or 0) >= 6 else \
                "MEDIUM" if (artifact.get("max_severity") or 0) >= 3.5 else "LOW"

    art_data = [
        ["Field", "Value"],
        ["Name",        artifact.get("name", "—")],
        ["Category",    artifact.get("category", "—")],
        ["Age",         f"{artifact.get('age', '—')} years"],
        ["Location",    artifact.get("location", "—")],
        ["Status",      artifact.get("status", "—")],
        ["Risk Level",  sev_label],
        ["Custodian",   artifact.get("custodian_name", "—")],
        ["Description", artifact.get("description", "—")],
    ]
    art_table = Table(art_data, colWidths=[4.5*cm, 12.5*cm])
    art_table.setStyle(TableStyle([
        ("BACKGROUND",  (0, 0), (-1, 0),  navy),
        ("TEXTCOLOR",   (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",    (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTNAME",    (0, 1), (0, -1),  "Helvetica-Bold"),
        ("BACKGROUND",  (0, 1), (0, -1),  colors.HexColor("#EBF1FA")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFF")]),
        ("GRID",        (0, 0), (-1, -1), 0.5, colors.HexColor("#C7D2FE")),
        ("PADDING",     (0, 0), (-1, -1), 7),
        ("FONTSIZE",    (0, 0), (-1, -1), 10),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(art_table)
    story.append(Spacer(1, 0.5*cm))

    # ── Inspection History ────────────────────────────
    if inspections:
        story.append(Paragraph("Inspection History", h2))
        insp_headers = ["Date", "Type", "Crack", "Fading", "Severity", "Risk", "Notes"]
        insp_data = [insp_headers]
        risk_row_colors = {
            "CRITICAL": colors.HexColor("#FEE2E2"),
            "HIGH":     colors.HexColor("#FFEDD5"),
            "MEDIUM":   colors.HexColor("#FEF9C3"),
            "LOW":      colors.HexColor("#DCFCE7"),
        }
        for insp in inspections[:30]:
            sev = float(insp.get("severity_index", 0))
            risk = ("CRITICAL" if sev >= 8 else "HIGH" if sev >= 6 else "MEDIUM" if sev >= 3.5 else "LOW")
            notes = (insp.get("damage_notes") or "")[:60] + ("…" if len(insp.get("damage_notes") or "") > 60 else "")
            insp_data.append([
                str(insp.get("inspection_date", ""))[:10],
                insp.get("inspection_type", "Routine"),
                "Yes" if insp.get("crack_detected") else "No",
                f"{float(insp.get('fading_level', 0)) * 100:.0f}%",
                f"{sev:.1f}",
                risk,
                notes,
            ])

        insp_table = Table(insp_data, colWidths=[2.5*cm, 3*cm, 1.8*cm, 2*cm, 2.2*cm, 2.5*cm, 4.5*cm])
        ts = TableStyle([
            ("BACKGROUND",  (0, 0), (-1, 0), navy),
            ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
            ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, -1), 9),
            ("GRID",        (0, 0), (-1, -1), 0.4, colors.HexColor("#C7D2FE")),
            ("PADDING",     (0, 0), (-1, -1), 6),
            ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFF")]),
        ])
        insp_table.setStyle(ts)
        story.append(insp_table)

    # ── Footer ────────────────────────────────────────
    story.append(Spacer(1, 1*cm))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.lightgrey))
    story.append(Spacer(1, 0.2*cm))
    footer_style = ParagraphStyle("AG_Footer", fontSize=9, textColor=colors.grey, alignment=TA_CENTER)
    story.append(Paragraph(
        f"Artifact Guardian AI System — Confidential Report — {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        footer_style))

    doc.build(story)
    buf.seek(0)
    return buf.getvalue()


# ── CSV Import ────────────────────────────────────────────────────────────────

def import_artifacts_from_csv(file_bytes: bytes) -> list:
    text = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    artifacts = []
    for row in reader:
        name = row.get("name", "").strip()
        if not name:
            continue
        
        artifacts.append({
            "name":        name,
            "category":    row.get("category", "Other").strip(),
            "age":         int(row.get("age", 0) or 0),
            "location":    row.get("location", "").strip(),
            "description": row.get("description", "").strip(),
            "image_url":   row.get("image_url", "").strip(),  # ← ADD THIS LINE

        })
    return artifacts
